from pathlib import Path
import traceback

REQUIRED_PACKAGES = ["openpyxl"]

EPW_DATA_HEADERS = [
    "Year", "Month", "Day", "Hour", "Minute",
    "Data Source and Uncertainty Flags",
    "Dry Bulb Temperature", "Dew Point Temperature", "Relative Humidity",
    "Atmospheric Station Pressure",
    "Extraterrestrial Horizontal Radiation",
    "Extraterrestrial Direct Normal Radiation",
    "Horizontal Infrared Radiation Intensity",
    "Global Horizontal Radiation",
    "Direct Normal Radiation",
    "Diffuse Horizontal Radiation",
    "Global Horizontal Illuminance",
    "Direct Normal Illuminance",
    "Diffuse Horizontal Illuminance",
    "Zenith Luminance",
    "Wind Direction", "Wind Speed",
    "Total Sky Cover", "Opaque Sky Cover",
    "Visibility", "Ceiling Height",
    "Present Weather Observation", "Present Weather Codes",
    "Precipitable Water", "Aerosol Optical Depth",
    "Snow Depth", "Days Since Last Snowfall",
    "Albedo",
    "Liquid Precipitation Depth", "Liquid Precipitation Quantity"
]

EPW_DATA_UNITS = [
    "[-]", "[-]", "[-]", "[-]", "[-]",
    "[-]",
    "[°C]", "[°C]", "[%]",
    "[Pa]",
    "[Wh/m²]",
    "[Wh/m²]",
    "[Wh/m²]",
    "[Wh/m²]",
    "[Wh/m²]",
    "[Wh/m²]",
    "[lux]",
    "[lux]",
    "[lux]",
    "[cd/m²]",
    "[degrees]", "[m/s]",
    "[tenths]", "[tenths]",
    "[km]", "[m]",
    "0= use weather code\n9=weather code missing",
    "TMY2 kód",
    "[mm]",
    "0.01 < clear, 0.015 = typical, 0.4 >high, 0.5 > extreme",
    "[cm]", "[days]",
    "[(W/m²)\n/(W/m²)",
    "[mm]", "[hr]"
]

SECTION_ORDER = [
    "LOCATION",
    "DESIGN CONDITIONS",
    "TYPICAL/EXTREME PERIODS",
    "GROUND TEMPERATURES",
    "HOLIDAYS/DAYLIGHT SAVINGS",
    "COMMENTS 1",
    "COMMENTS 2",
    "DATA PERIODS"
]

def pause():
    input("\nPress Enter to close...")

def check_packages():
    missing = []
    for pkg in REQUIRED_PACKAGES:
        try:
            __import__(pkg)
        except Exception:
            missing.append(pkg)
    if missing:
        print("Missing required packages:", ", ".join(missing))
        print("Install them with:")
        print(f"python -m pip install {' '.join(missing)}")
        return False
    return True

def split_csv_line(line):
    return [x.strip() for x in line.split(",")]

def wrap_long_text(value):
    if value is None:
        return None
    text = str(value)
    if " " in text and len(text) > 10:
        return "\n".join(text.split())
    return value

def try_number(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        if "." in s or "e" in s.lower():
            return float(s)
        return int(s)
    except ValueError:
        return wrap_long_text(s)

def read_epw_lines(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        return [line.rstrip("\n\r") for line in f]

def month_name_from_token(token):
    t = token.strip()
    if "=" in t:
        left, right = t.split("=", 1)
        return left.strip(), right.strip()
    return t, ""

def write_cell(ws, row, col, value, bold=False, wrap=True):
    from openpyxl.styles import Font, Alignment
    c = ws.cell(row=row, column=col, value=value)
    if bold:
        c.font = Font(bold=True)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

def write_table(ws, start_row, start_col, rows, bold_first_row=False):
    for r_offset, row in enumerate(rows):
        for c_offset, value in enumerate(row):
            write_cell(
                ws,
                start_row + r_offset,
                start_col + c_offset,
                value,
                bold=(bold_first_row and r_offset == 0),
                wrap=True
            )

def build_section_tables(header_lines):
    tables = []

    loc = split_csv_line(header_lines[0])
    tables.append((
        "LOCATION",
        [
            ["Field", "Value 1", "Value 2", "Value 3", "Value 4", "Value 5", "Value 6", "Value 7", "Value 8", "Value 9"],
            ["City", loc[1] if len(loc) > 1 else ""],
            ["State/Province", loc[2] if len(loc) > 2 else ""],
            ["Country", loc[3] if len(loc) > 3 else ""],
            ["Source", loc[4] if len(loc) > 4 else ""],
            ["WMO", loc[5] if len(loc) > 5 else ""],
            ["Latitude", loc[6] if len(loc) > 6 else ""],
            ["Longitude", loc[7] if len(loc) > 7 else ""],
            ["Time Zone", loc[8] if len(loc) > 8 else ""],
            ["Elevation", loc[9] if len(loc) > 9 else ""],
        ]
    ))

    for title, line_idx in [
        ("DESIGN CONDITIONS", 1),
        ("GROUND TEMPERATURES", 3),
        ("HOLIDAYS/DAYLIGHT SAVINGS", 4),
        ("COMMENTS 1", 5),
        ("COMMENTS 2", 6),
    ]:
        parts = split_csv_line(header_lines[line_idx])
        rows = [["Field"] + [f"Value {i}" for i in range(1, max(2, len(parts)))]]
        rows.append(["Raw Line"] + parts[1:])
        tables.append((title, rows))

    typ = split_csv_line(header_lines[2])
    typ_rows = [["Field", "Value 1", "Value 2", "Value 3", "Value 4"]]
    if len(typ) > 1:
        typ_rows.append(["Number of Typical/Extreme Periods", typ[1]])
    idx = 2
    p = 1
    while idx + 3 < len(typ):
        typ_rows.append([
            f"Period {p}",
            typ[idx],
            typ[idx + 1],
            typ[idx + 2],
            typ[idx + 3],
        ])
        idx += 4
        p += 1
    tables.append(("TYPICAL/EXTREME PERIODS", typ_rows))

    dp = split_csv_line(header_lines[7])
    dp_rows = [["Year", "Month", "Value 3", "Value 4", "Value 5", "Value 6", "Value 7"]]
    if len(dp) > 1:
        idx = 1
        while idx < len(dp):
            token = dp[idx]
            if "=" in token:
                month, year = month_name_from_token(token)
                row = [year, month]
                for extra in dp[idx + 1:idx + 5]:
                    row.append(extra)
                while len(row) < 7:
                    row.append("")
                dp_rows.append(row[:7])
                idx += 5
            else:
                dp_rows.append([token])
                idx += 1
    tables.append(("DATA PERIODS", dp_rows))

    return tables

def autofit(ws, max_width=55):
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            txt = str(cell.value)
            for line in txt.split("\n"):
                widths[cell.column] = max(widths.get(cell.column, 0), len(line))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), max_width)

def main():
    try:
        if not check_packages():
            pause()
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        source = input("Enter the full path of the EPW file: ").strip().strip('"')
        source_path = Path(source)

        if not source_path.exists():
            print("Source file not found.")
            pause()
            return

        lines = read_epw_lines(source_path)
        if len(lines) < 8:
            print("The file does not look like a valid EPW file.")
            pause()
            return

        header_lines = lines[:8]
        data_lines = lines[8:]

        wb = Workbook()
        ws = wb.active
        ws.title = "EPW Header"
        report_ws = wb.create_sheet("Report")

        row = 1
        section_tables = build_section_tables(header_lines)

        for section_title, table in section_tables:
            write_cell(ws, row, 1, section_title, bold=True)
            row += 1
            write_table(ws, row, 1, table, bold_first_row=True)
            row += len(table) + 2

        ws2 = wb.create_sheet("EPW Data")
        write_table(ws2, 1, 1, [EPW_DATA_HEADERS, EPW_DATA_UNITS], bold_first_row=True)

        for r_idx, line in enumerate(data_lines, start=3):
            parts = split_csv_line(line)
            for c_idx, val in enumerate(parts, start=1):
                write_cell(ws2, r_idx, c_idx, try_number(val), wrap=True)

        report = [
            "Summary Report",
            f"Source file: {source_path}",
            f"Output file: {source_path.with_suffix('.xlsx')}",
            f"Header records processed: {len(header_lines)}",
            f"Hourly data rows processed: {len(data_lines)}",
            "Header sections written as separate tables.",
            "Data periods split into year and month columns when possible.",
            "Cells containing spaces and longer than 10 characters were wrapped.",
        ]

        for i, txt in enumerate(report, start=1):
            write_cell(report_ws, i, 1, txt)
        report_ws.column_dimensions["A"].width = 120

        autofit(ws)
        autofit(ws2)
        output_path = source_path.with_suffix(".xlsx")
        wb.save(output_path)

        print(f"Saved: {output_path}")
        for line in report[1:]:
            print(line)

    except Exception as e:
        print("Error occurred:")
        print(e)
        traceback.print_exc()
    finally:
        pause()

if __name__ == "__main__":
    main()